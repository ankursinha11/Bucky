"""
Excel Exporter
Exports STTM reports and gap analysis to Excel files
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from loguru import logger

from core.models import STTMReport, Gap, GapSeverity


class ExcelExporter:
    """Export reports to Excel format"""

    def __init__(self, output_dir: str = "./outputs/reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_sttm_report(self, report: STTMReport, filename: Optional[str] = None) -> str:
        """Export STTM report to Excel"""
        if not filename:
            filename = f"STTM_{report.report_name.replace(' ', '_')}.xlsx"

        output_path = self.output_dir / filename
        logger.info(f"Exporting STTM report to {output_path}")

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: Summary
        self._create_sttm_summary_sheet(wb, report)

        # Sheet 2: Mappings
        self._create_sttm_mappings_sheet(wb, report)

        # Sheet 3: Statistics
        self._create_sttm_statistics_sheet(wb, report)

        wb.save(output_path)
        logger.info(f"STTM report exported successfully: {output_path}")
        return str(output_path)

    def export_gap_analysis(
        self, gaps: List[Gap], filename: str = "Gap_Analysis_Report.xlsx"
    ) -> str:
        """Export gap analysis to Excel"""
        output_path = self.output_dir / filename
        logger.info(f"Exporting gap analysis to {output_path}")

        wb = Workbook()
        wb.remove(wb.active)

        # Sheet 1: Summary
        self._create_gap_summary_sheet(wb, gaps)

        # Sheet 2: All Gaps
        self._create_gaps_detail_sheet(wb, gaps, "All Gaps")

        # Sheet 3: Critical Gaps
        critical_gaps = [g for g in gaps if g.severity == GapSeverity.CRITICAL]
        if critical_gaps:
            self._create_gaps_detail_sheet(wb, critical_gaps, "Critical Gaps")

        # Sheet 4: By Type
        self._create_gaps_by_type_sheets(wb, gaps)

        wb.save(output_path)
        logger.info(f"Gap analysis exported successfully: {output_path}")
        return str(output_path)

    def export_combined_report(
        self, sttm_report: STTMReport, gaps: List[Gap], filename: str = "Combined_Analysis_Report.xlsx"
    ) -> str:
        """Export combined STTM and gap analysis report"""
        output_path = self.output_dir / filename
        logger.info(f"Exporting combined report to {output_path}")

        wb = Workbook()
        wb.remove(wb.active)

        # Executive Summary
        self._create_executive_summary_sheet(wb, sttm_report, gaps)

        # STTM sheets
        self._create_sttm_mappings_sheet(wb, sttm_report)

        # Gap sheets
        self._create_gaps_detail_sheet(wb, gaps, "Gap Analysis")

        wb.save(output_path)
        logger.info(f"Combined report exported successfully: {output_path}")
        return str(output_path)

    def _create_sttm_summary_sheet(self, wb: Workbook, report: STTMReport):
        """Create STTM summary sheet"""
        ws = wb.create_sheet("Summary")

        # Title
        ws["A1"] = "Source-to-Target Mapping Report"
        ws["A1"].font = Font(size=16, bold=True)

        summary_data = report.get_summary()

        row = 3
        for key, value in summary_data.items():
            ws[f"A{row}"] = str(key).replace("_", " ").title()
            ws[f"B{row}"] = str(value)
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Auto-width
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50

    def _create_sttm_mappings_sheet(self, wb: Workbook, report: STTMReport):
        """Create STTM mappings sheet"""
        ws = wb.create_sheet("Source_To_Target_Mapping")

        # Convert to DataFrame
        df = report.to_dataframe()

        # Write to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Style header
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center")

        # Auto-width columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_sttm_statistics_sheet(self, wb: Workbook, report: STTMReport):
        """Create statistics sheet"""
        ws = wb.create_sheet("Statistics")

        stats = [
            ["Metric", "Value"],
            ["Total Mappings", report.total_mappings],
            ["Direct Mappings", report.direct_mappings],
            ["Derived Mappings", report.derived_mappings],
            ["Lookup Mappings", report.lookup_mappings],
            ["Calculated Mappings", report.calculated_mappings],
            ["Source Tables", len(report.source_tables)],
            ["Target Tables", len(report.target_tables)],
        ]

        for r_idx, row in enumerate(stats, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _create_gap_summary_sheet(self, wb: Workbook, gaps: List[Gap]):
        """Create gap summary sheet"""
        ws = wb.create_sheet("Gap Summary")

        # Title
        ws["A1"] = "Gap Analysis Summary"
        ws["A1"].font = Font(size=16, bold=True)

        # Count by type
        from collections import Counter

        by_type = Counter([g.gap_type.value for g in gaps])
        by_severity = Counter([g.severity.value for g in gaps])

        row = 3
        ws[f"A{row}"] = "Total Gaps"
        ws[f"B{row}"] = len(gaps)
        ws[f"A{row}"].font = Font(bold=True)
        row += 2

        # By Type
        ws[f"A{row}"] = "Gaps by Type"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        for gap_type, count in by_type.items():
            ws[f"A{row}"] = gap_type
            ws[f"B{row}"] = count
            row += 1

        row += 1
        # By Severity
        ws[f"A{row}"] = "Gaps by Severity"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        for severity, count in by_severity.items():
            ws[f"A{row}"] = severity
            ws[f"B{row}"] = count
            # Color code
            if severity == "critical":
                ws[f"A{row}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            elif severity == "high":
                ws[f"A{row}"].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            row += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15

    def _create_gaps_detail_sheet(self, wb: Workbook, gaps: List[Gap], sheet_name: str):
        """Create detailed gaps sheet"""
        ws = wb.create_sheet(sheet_name)

        # Convert to DataFrame
        gap_data = []
        for gap in gaps:
            gap_data.append(
                {
                    "Gap ID": gap.gap_id,
                    "Type": gap.gap_type.value,
                    "Severity": gap.severity.value,
                    "Title": gap.title,
                    "Source System": gap.source_system,
                    "Target System": gap.target_system,
                    "Source Process": gap.source_process_name or "",
                    "Target Process": gap.target_process_name or "MISSING",
                    "Description": gap.description,
                    "Impact": gap.business_impact or "",
                    "Recommendation": gap.recommendation or "",
                    "Confidence": gap.confidence_score,
                }
            )

        df = pd.DataFrame(gap_data)

        # Write to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Style header
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                else:
                    # Color code by severity
                    severity = df.iloc[r_idx - 2]["Severity"]
                    if severity == "critical":
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif severity == "high":
                        cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")

        # Auto-width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

        ws.freeze_panes = "A2"

    def _create_gaps_by_type_sheets(self, wb: Workbook, gaps: List[Gap]):
        """Create sheets for each gap type"""
        from collections import defaultdict

        gaps_by_type = defaultdict(list)
        for gap in gaps:
            gaps_by_type[gap.gap_type.value].append(gap)

        for gap_type, type_gaps in gaps_by_type.items():
            sheet_name = f"{gap_type[:25]}"  # Limit sheet name length
            self._create_gaps_detail_sheet(wb, type_gaps, sheet_name)

    def _create_executive_summary_sheet(
        self, wb: Workbook, sttm_report: STTMReport, gaps: List[Gap]
    ):
        """Create executive summary"""
        ws = wb.create_sheet("Executive Summary", 0)  # Insert as first sheet

        ws["A1"] = "Codebase Intelligence - Executive Summary"
        ws["A1"].font = Font(size=18, bold=True)

        row = 3

        # STTM Summary
        ws[f"A{row}"] = "Source-to-Target Mapping"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        sttm_summary = sttm_report.get_summary()
        ws[f"A{row}"] = "Total Mappings"
        ws[f"B{row}"] = sttm_summary["total_mappings"]
        row += 1
        ws[f"A{row}"] = "Source Tables"
        ws[f"B{row}"] = sttm_summary["source_tables_count"]
        row += 1
        ws[f"A{row}"] = "Target Tables"
        ws[f"B{row}"] = sttm_summary["target_tables_count"]
        row += 2

        # Gap Analysis Summary
        ws[f"A{row}"] = "Gap Analysis"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        ws[f"A{row}"] = "Total Gaps Identified"
        ws[f"B{row}"] = len(gaps)
        row += 1

        from collections import Counter

        by_severity = Counter([g.severity.value for g in gaps])
        for severity, count in by_severity.items():
            ws[f"A{row}"] = f"{severity.title()} Severity"
            ws[f"B{row}"] = count
            row += 1

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20


from typing import Optional  # Add missing import
