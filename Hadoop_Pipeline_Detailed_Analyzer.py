#!/usr/bin/env python3
"""
Enhanced Hadoop Pipeline Analyzer
- Automatically finds all repos in a folder
- Analyzes each workflow.xml file individually
- Shows script details for each pipeline
"""

import os
from pathlib import Path
from typing import Dict, List
import json
from datetime import datetime


class DetailedPipelineAnalyzer:
    """Detailed analysis at the individual workflow.xml level"""

    def __init__(self):
        self.results = []

    def analyze_folder(self, folder_path: str):
        """Analyze all repos in a folder"""
        folder = Path(folder_path)
        
        # Find all repos (directories with workflows/)
        repos = []
        for item in folder.iterdir():
            if item.is_dir():
                workflows_dir = item / "workflows"
                if workflows_dir.exists():
                    repos.append(item)

        print(f"📂 Found {len(repos)} repositories in: {folder_path}\n")

        for repo in repos:
            self.analyze_repo(str(repo))

    def analyze_repo(self, repo_path: str):
        """Analyze a single repository"""
        repo = Path(repo_path)
        repo_name = repo.name

        print(f"\n{'='*80}")
        print(f"🔍 Repository: {repo_name}")
        print(f"{'='*80}")

        # Find all workflow.xml files
        workflows_dir = repo / "workflows"
        if not workflows_dir.exists():
            print(f"⚠️ No workflows directory found")
            return

        # Find all XML files (workflows and coordinators)
        xml_files = list(repo.rglob("*.xml"))
        
        print(f"📋 Analyzing individual workflow.xml files...\n")

        for xml_file in xml_files:
            # Skip if it's in a deep subdirectory we don't care about
            if "workflows" not in str(xml_file):
                continue

            # Determine type
            is_coordinator = "coordinator" in xml_file.name.lower()
            
            if is_coordinator:
                continue  # Skip coordinators for now

            # Get workflow name from path
            workflow_name = self._extract_workflow_name(xml_file)
            
            print(f"  📄 Pipeline: {xml_file.name}")
            print(f"     Location: {workflow_name}")

            # Analyze this workflow.xml
            pipeline_data = self._analyze_workflow_xml(xml_file)
            
            # Find associated scripts
            scripts_data = self._find_associated_scripts(xml_file)
            
            # Combine
            result = {
                'repo': repo_name,
                'pipeline_name': xml_file.name,
                'workflow_location': workflow_name,
                'full_path': str(xml_file),
                'actions': pipeline_data['actions'],
                'tables': pipeline_data.get('tables', []),
                'table_count': pipeline_data.get('table_count', 0),
                'scripts': scripts_data
            }
            
            self.results.append(result)
            
            # Print summary
            total_files = len(scripts_data)
            total_loc = sum(s.get('lines', 0) for s in scripts_data.values())
            size = self._classify_size(total_files, total_loc, pipeline_data['actions'], pipeline_data.get('table_count', 0))
            
            print(f"     ⚙️  Actions: {pipeline_data['actions']}")
            print(f"     📊 Tables: {pipeline_data.get('table_count', 0)}")
            print(f"     📄 Files: {total_files} ({self._format_file_counts(scripts_data)})")
            print(f"     📝 LOC: {total_loc:,}")
            print(f"     {size}")
            print()

    def _extract_workflow_name(self, xml_file: Path) -> str:
        """Extract workflow location from path"""
        parts = xml_file.parts
        if "workflows" in parts:
            idx = parts.index("workflows")
            if idx + 1 < len(parts):
                return "/".join(parts[idx:idx+2])
        return str(xml_file)

    def _analyze_workflow_xml(self, xml_file: Path) -> Dict:
        """Analyze a single workflow.xml file"""
        try:
            import xml.etree.ElementTree as ET
            import re
            
            tree = ET.parse(xml_file)
            root = tree.getroot()
            
            actions = root.findall('.//{uri:oozie:workflow:0.5}action')
            
            # Extract tables/datasets from the workflow
            all_text = ET.tostring(root).decode('utf-8', errors='ignore')
            
            # Find common table patterns
            hdfs_patterns = re.findall(r'hdfs://[^<>"\s]+', all_text)
            tables = []
            for pattern in hdfs_patterns:
                # Extract table name from path (last segment)
                if '/' in pattern:
                    segments = pattern.split('/')
                    # Skip common Hadoop paths
                    if 'warehouse' in segments or 'user' in segments or 'tmp' in segments or 'stage' in segments:
                        continue
                    # Get meaningful table names
                    table_name = segments[-1] if segments else ''
                    if table_name and len(table_name) > 1 and table_name not in tables:
                        tables.append(table_name)
            
            # Also look in configuration properties
            config_elements = root.findall('.//{uri:oozie:workflow:0.5}configuration')
            table_mentions = []
            for config in config_elements:
                config_text = ET.tostring(config).decode('utf-8', errors='ignore')
                # Look for common table identifiers
                table_refs = re.findall(r'(?:\w+\.\w+\.\w+|/\w+/\w+)/(\w+)', config_text)
                table_mentions.extend(table_refs)
            
            unique_tables = list(set(tables + table_mentions))
            
            return {
                'actions': len(actions),
                'tables': unique_tables,
                'table_count': len(unique_tables)
            }
        except Exception as e:
            return {'actions': 0, 'tables': [], 'table_count': 0}

    def _find_associated_scripts(self, workflow_xml: Path) -> Dict:
        """Find all scripts associated with this workflow"""
        scripts = {}
        
        # Get the base directory (parent of oozie/)
        base_dir = workflow_xml.parent.parent
        
        # PySpark scripts
        spark_dir = base_dir / "spark"
        if spark_dir.exists():
            for py_file in spark_dir.glob("*.py"):
                lines = self._count_lines(py_file)
                scripts[str(py_file.name)] = {
                    'type': 'PySpark',
                    'path': str(py_file),
                    'lines': lines
                }

        # Pig scripts
        pig_dir = base_dir / "pig"
        if pig_dir.exists():
            for pig_file in pig_dir.glob("*.pig"):
                lines = self._count_lines(pig_file)
                scripts[str(pig_file.name)] = {
                    'type': 'Pig',
                    'path': str(pig_file),
                    'lines': lines
                }

        # Shell scripts
        shell_dir = base_dir / "shell"
        if shell_dir.exists():
            for sh_file in shell_dir.glob("*.sh"):
                lines = self._count_lines(sh_file)
                scripts[str(sh_file.name)] = {
                    'type': 'Shell',
                    'path': str(sh_file),
                    'lines': lines
                }

        # Hive SQL
        hive_dir = base_dir / "hive"
        if hive_dir.exists():
            for sql_file in list(hive_dir.glob("*.sql")) + list(hive_dir.glob("*.hql")):
                lines = self._count_lines(sql_file)
                scripts[str(sql_file.name)] = {
                    'type': 'Hive',
                    'path': str(sql_file),
                    'lines': lines
                }

        return scripts

    def _count_lines(self, file_path: Path) -> int:
        """Count lines in a file"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return len(f.readlines())
        except:
            return 0

    def _format_file_counts(self, scripts: Dict) -> str:
        """Format file counts by type"""
        counts = {}
        for script_data in scripts.values():
            script_type = script_data['type']
            counts[script_type] = counts.get(script_type, 0) + 1
        
        parts = [f"{v} {k}" for k, v in counts.items()]
        return ", ".join(parts)

    def _classify_size(self, files: int, loc: int, actions: int, tables: int) -> str:
        """Classify pipeline size based on multiple factors"""
        # Calculate complexity score
        # File count score
        file_score = min(files / 2, 1.0)
        # LOC score  
        loc_score = min(loc / 2000, 1.0)
        # Actions score
        action_score = min(actions / 10, 1.0)
        # Table count score
        table_score = min(tables / 5, 1.0)
        
        # Overall complexity
        complexity = (file_score + loc_score + action_score + table_score) / 4
        
        if complexity <= 0.25:
            return "🟢 Size: SMALL"
        elif complexity <= 0.50:
            return "🟡 Size: MEDIUM"
        elif complexity <= 0.75:
            return "🟠 Size: LARGE"
        else:
            return "🔴 Size: XLARGE"

    def export_to_excel(self, output_file: str = None):
        """Export results to Excel"""
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"Hadoop_Pipeline_Detailed_Analysis_{timestamp}.xlsx"

        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = Workbook()

            # Sheet 1: Pipeline Summary
            ws1 = wb.active
            ws1.title = 'Pipeline Summary'

            headers1 = ['Repo', 'Pipeline', 'Location', 'Actions', 'Tables', 'Total Files', 'Total LOC', 'Size']
            ws1.append(headers1)

            # Style headers
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            for cell in ws1[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Add data
            for r in self.results:
                total_files = len(r['scripts'])
                total_loc = sum(s.get('lines', 0) for s in r['scripts'].values())
                size = self._classify_size(total_files, total_loc, r['actions'], r.get('table_count', 0)).split(': ')[1]
                
                ws1.append([
                    r['repo'],
                    r['pipeline_name'],
                    r['workflow_location'],
                    r['actions'],
                    r.get('table_count', 0),
                    total_files,
                    total_loc,
                    size
                ])

            # Auto-adjust widths
            for col in range(1, len(headers1) + 1):
                ws1.column_dimensions[get_column_letter(col)].width = 25

            # Sheet 2: Detailed Script List
            ws2 = wb.create_sheet('Script Details')

            headers2 = ['Repo', 'Pipeline', 'Script', 'Type', 'Lines']
            ws2.append(headers2)

            # Style headers
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Add data
            for r in self.results:
                for script_name, script_data in r['scripts'].items():
                    ws2.append([
                        r['repo'],
                        r['pipeline_name'],
                        script_name,
                        script_data['type'],
                        script_data.get('lines', 0)
                    ])

            # Auto-adjust widths
            for col in range(1, len(headers2) + 1):
                ws2.column_dimensions[get_column_letter(col)].width = 30

            wb.save(output_file)
            print(f"\n✅ Excel exported to: {output_file}")

        except ImportError:
            print("⚠️  pandas/openpyxl not installed")
            # Save JSON instead
            json_file = output_file.replace('.xlsx', '.json')
            with open(json_file, 'w') as f:
                json.dump(self.results, f, indent=2, default=str)
            print(f"✅ JSON exported to: {json_file}")


def main():
    """Main function"""
    import argparse

    parser = argparse.ArgumentParser(
        description='Analyze Hadoop repositories at the individual pipeline level'
    )

    parser.add_argument(
        'path',
        help='Path to Hadoop repository or folder containing multiple repos'
    )

    parser.add_argument(
        '--output',
        type=str,
        default=None,
        help='Output file name'
    )

    args = parser.parse_args()

    analyzer = DetailedPipelineAnalyzer()

    # Check if it's a single repo or folder
    if os.path.isdir(args.path):
        workflows_dir = Path(args.path) / "workflows"
        if workflows_dir.exists():
            # Single repo
            print("Single repository detected")
            analyzer.analyze_repo(args.path)
        else:
            # Folder of repos
            print("Folder of repositories detected")
            analyzer.analyze_folder(args.path)
    else:
        print(f"⚠️ Not a valid directory: {args.path}")

    # Export
    analyzer.export_to_excel(args.output)

    print("\n" + "="*80)
    print("✅ Analysis Complete!")
    print("="*80)


if __name__ == "__main__":
    main()

