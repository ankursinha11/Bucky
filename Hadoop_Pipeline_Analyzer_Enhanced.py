#!/usr/bin/env python3
"""
Complete Hadoop Repository Analyzer - Two-Phase Approach

Phase 1: Scan entire repo and catalog ALL scripts (wherever they are)
Phase 2: Find all workflows/coordinators and match scripts

Handles all possible directory structures and patterns.
"""

import os
import re
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Dict, List, Set, Tuple
from datetime import datetime
from collections import defaultdict


class CompleteHadoopAnalyzer:
    """Two-phase analyzer for complete Hadoop repository analysis"""

    def __init__(self):
        self.repos_data = []  # All analyzed repos
        self.script_inventory = {}  # Complete script inventory per repo

    def analyze_folder(self, folder_path: str):
        """Analyze all Hadoop repos in a folder"""
        folder = Path(folder_path)
        
        print(f"📂 Scanning: {folder_path}\n")
        
        # Find all repos (directories with certain indicators)
        repos = []
        for item in folder.iterdir():
            if item.is_dir() and not item.name.startswith('.'):
                # Check if it looks like a Hadoop repo
                if self._is_hadoop_repo(item):
                    repos.append(item)
                    print(f"   ✓ Found repo: {item.name}")

        print(f"\n✅ Found {len(repos)} Hadoop repositories\n")

        for repo in repos:
            self.analyze_repo(str(repo))

    def _is_hadoop_repo(self, path: Path) -> bool:
        """Check if directory is a Hadoop repository"""
        # Look for indicators
        indicators = [
            (path / "workflows").exists(),
            (path / "oozie").exists(),
            len(list(path.rglob("*.xml"))) > 0,
            len(list(path.rglob("*.pig"))) > 0,
            len(list(path.rglob("*.py"))) > 0,
        ]
        return any(indicators)

    def analyze_repo(self, repo_path: str):
        """Analyze a single Hadoop repository - TWO PHASE"""
        repo = Path(repo_path)
        repo_name = repo.name

        print(f"\n{'='*100}")
        print(f"🔍 Repository: {repo_name}")
        print(f"{'='*100}")

        # ============= PHASE 1: SCAN ALL SCRIPTS =============
        print(f"\n📋 PHASE 1: Scanning all scripts in repository...")
        script_inventory = self._scan_all_scripts(repo)
        
        total_scripts = sum(len(scripts) for scripts in script_inventory.values())
        print(f"   ✅ Found {total_scripts} total scripts:")
        for script_type, scripts in script_inventory.items():
            if scripts:
                print(f"      - {script_type}: {len(scripts)} files")
        
        # Store inventory
        self.script_inventory[repo_name] = script_inventory

        # ============= PHASE 2: FIND AND ANALYZE WORKFLOWS =============
        print(f"\n📋 PHASE 2: Finding and analyzing workflows/coordinators...")
        workflows = self._find_all_workflows(repo)
        coordinators = self._find_all_coordinators(repo)
        
        print(f"   ✅ Found {len(workflows)} unique workflows")
        print(f"   ✅ Found {len(coordinators)} coordinators")
        
        # Show workflow list for verification
        if len(workflows) <= 20:
            print(f"\n   Workflows found:")
            for wf in workflows:
                print(f"      - {wf.relative_to(repo)}")
        print()

        # Analyze each workflow
        for workflow_file in workflows:
            self._analyze_workflow(repo, repo_name, workflow_file, script_inventory)

    def _scan_all_scripts(self, repo: Path) -> Dict[str, List[Dict]]:
        """Phase 1: Scan entire repo and catalog ALL scripts"""
        script_inventory = {
            'PySpark': [],
            'Pig': [],
            'Shell': [],
            'Hive': [],
            'Java': [],
            'MapReduce': [],
            'Other': []
        }
        
        # Define patterns
        patterns = {
            'PySpark': '*.py',
            'Pig': '*.pig',
            'Shell': '*.sh',
            'Hive': ['*.sql', '*.hql'],
            'Java': '*.jar',
        }
        
        # Search entire repo recursively
        for script_type, pattern_list in patterns.items():
            if isinstance(pattern_list, str):
                pattern_list = [pattern_list]
            
            for pattern in pattern_list:
                for script_file in repo.rglob(pattern):
                    # Skip certain directories
                    if any(skip in str(script_file) for skip in ['.git', '__pycache__', 'target', 'build']):
                        continue
                    
                    lines = self._count_lines(script_file)
                    rel_path = script_file.relative_to(repo)
                    
                    script_inventory[script_type].append({
                        'name': script_file.name,
                        'path': str(script_file),
                        'relative_path': str(rel_path),
                        'lines': lines,
                        'used': False  # Will be marked True if referenced in workflow
                    })
        
        return script_inventory

    def _find_all_workflows(self, repo: Path) -> List[Path]:
        """Find all workflow XML files (not coordinators)"""
        workflows = []
        seen_names = set()  # Track unique workflow names to avoid duplicates
        
        # Search for all XML files
        for xml_file in repo.rglob("*.xml"):
            # Skip certain directories
            if any(skip in str(xml_file) for skip in ['.git', 'target', 'build', 'target', '.svn']):
                continue
            
            # Skip backup/temp files
            if any(xml_file.name.endswith(suffix) for suffix in ['.bak', '.tmp', '.backup', '~']):
                continue
            
            # Check if it's a workflow (not coordinator)
            try:
                tree = ET.parse(xml_file)
                root = tree.getroot()
                
                # Check root tag
                root_tag = root.tag.split('}')[-1] if '}' in root.tag else root.tag
                
                if root_tag == 'workflow-app':
                    # Get workflow name to avoid duplicates
                    workflow_name = root.attrib.get('name', '')
                    
                    # Create unique key based on name or file path
                    if workflow_name:
                        unique_key = workflow_name
                    else:
                        # Use relative path as key
                        unique_key = str(xml_file.relative_to(repo))
                    
                    # Only add if we haven't seen this workflow name before
                    if unique_key not in seen_names:
                        workflows.append(xml_file)
                        seen_names.add(unique_key)
            except:
                continue
        
        return workflows

    def _find_all_coordinators(self, repo: Path) -> List[Path]:
        """Find all coordinator XML files"""
        coordinators = []
        
        # Search for all XML files
        for xml_file in repo.rglob("*.xml"):
            # Skip certain directories
            if any(skip in str(xml_file) for skip in ['.git', 'target', 'build']):
                continue
            
            # Check if it's a coordinator
            try:
                tree = ET.parse(xml_file)
                root = tree.getroot()
                
                # Check root tag
                root_tag = root.tag.split('}')[-1] if '}' in root.tag else root.tag
                
                if root_tag == 'coordinator-app':
                    coordinators.append(xml_file)
            except:
                continue
        
        return coordinators

    def _analyze_workflow(self, repo: Path, repo_name: str, workflow_file: Path, script_inventory: Dict):
        """Analyze a single workflow file"""
        
        # Extract workflow info
        workflow_name = self._extract_workflow_name(workflow_file)
        module = self._extract_module_name(repo, workflow_file)
        
        print(f"  📄 {workflow_name}")
        print(f"     Module: {module}")
        print(f"     Path: {workflow_file.relative_to(repo)}")
        
        # Parse workflow XML
        workflow_data = self._parse_workflow_xml(workflow_file)
        
        # Match referenced scripts with inventory
        referenced_scripts = workflow_data['referenced_scripts']
        matched_scripts, missing_scripts = self._match_scripts(referenced_scripts, script_inventory)
        
        # Mark scripts as used
        for script_name in matched_scripts:
            self._mark_script_used(script_name, script_inventory)
        
        # Calculate metrics
        total_lines = sum(s['lines'] for s in matched_scripts.values())
        
        # Classify size
        pipeline_length = self._classify_pipeline_length(
            len(referenced_scripts),
            total_lines,
            workflow_data['actions'],
            workflow_data['table_count']
        )
        
        print(f"     📊 Actions: {workflow_data['actions']}, Tables: {workflow_data['table_count']}")
        print(f"     📝 Scripts: {len(referenced_scripts)} ({len(matched_scripts)} found, {len(missing_scripts)} missing)")
        print(f"     📏 LOC: {total_lines:,}")
        print(f"     {pipeline_length}")
        
        if missing_scripts:
            print(f"     ⚠️  Missing: {', '.join(list(missing_scripts)[:5])}")
            if len(missing_scripts) > 5:
                print(f"              ... and {len(missing_scripts) - 5} more")
        print()
        
        # Calculate effort estimation
        effort = self._estimate_effort(
            scripts=len(referenced_scripts),
            loc=total_lines,
            actions=workflow_data['actions'],
            tables=workflow_data['table_count'],
            pipeline_size=pipeline_length
        )
        
        # Store result
        self.repos_data.append({
            'repo': repo_name,
            'pipeline_name': workflow_name,
            'module': module,
            'workflow_path': str(workflow_file.relative_to(repo)),
            'no_of_scripts': len(referenced_scripts),
            'scripts_found': len(matched_scripts),
            'scripts_missing': len(missing_scripts),
            'pipeline_length': pipeline_length,
            'actions': workflow_data['actions'],
            'tables': workflow_data['table_count'],
            'total_loc': total_lines,
            'effort_days': effort['total_days'],
            'effort_hours': effort['total_hours'],
            'effort_estimate': effort['effort_estimate'],
            'matched_scripts': matched_scripts,
            'missing_scripts': list(missing_scripts),
            'referenced_scripts': list(referenced_scripts),
            'effort_breakdown': effort  # Full breakdown for detailed sheet
        })

    def _extract_workflow_name(self, workflow_file: Path) -> str:
        """Extract workflow name from XML name attribute or path"""
        try:
            tree = ET.parse(workflow_file)
            root = tree.getroot()
            
            if 'name' in root.attrib:
                xml_name = root.attrib['name']
                # Clean up the name
                if xml_name:
                    # Remove common prefixes
                    cleaned = xml_name.replace('escan_data_ingestion', '').replace(':', '').strip()
                    # Remove leading/trailing special chars
                    cleaned = re.sub(r'^[\s:_-]+|[\s:_-]+$', '', cleaned)
                    if cleaned:
                        return cleaned
        except:
            pass
        
        # Fallback: use filename or parent directory
        name = workflow_file.stem
        if name == 'workflow':
            # Try parent directory
            parent = workflow_file.parent.name
            if parent != 'oozie':
                return parent
        
        return name

    def _extract_module_name(self, repo: Path, workflow_file: Path) -> str:
        """Extract module name from workflow path"""
        try:
            rel_path = workflow_file.relative_to(repo)
            parts = rel_path.parts
            
            # Look for common module indicators
            for i, part in enumerate(parts):
                if part in ['workflows', 'oozie'] and i + 1 < len(parts):
                    return parts[i + 1]
            
            # Fallback: use parent directory
            return workflow_file.parent.name
        except:
            return workflow_file.parent.name

    def _parse_workflow_xml(self, workflow_file: Path) -> Dict:
        """Parse workflow XML to extract actions, tables, and script references"""
        try:
            tree = ET.parse(workflow_file)
            root = tree.getroot()
            
            # Get all text content
            all_text = ET.tostring(root, encoding='unicode')
            
            # Count actions
            action_count = 0
            for elem in root.iter():
                tag_name = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
                if tag_name == 'action':
                    action_count += 1
            
            # Extract referenced scripts
            referenced_scripts = set()
            
            for elem in root.iter():
                tag_name = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
                
                # Look for script-related tags
                if tag_name in ['script', 'file', 'exec', 'jar', 'app-path', 'main-class']:
                    if elem.text:
                        script_path = elem.text.strip()
                        # Skip variables
                        if '$' in script_path or not script_path:
                            continue
                        
                        # Extract filename
                        filename = script_path.split('/')[-1] if '/' in script_path else script_path
                        
                        # Only include actual script files
                        if any(filename.endswith(ext) for ext in ['.py', '.pig', '.sh', '.sql', '.hql', '.jar']):
                            referenced_scripts.add(filename)
            
            # Extract tables
            tables = self._extract_tables_from_xml(all_text)
            
            return {
                'actions': action_count,
                'table_count': len(tables),
                'tables': tables,
                'referenced_scripts': referenced_scripts
            }
            
        except Exception as e:
            print(f"     ⚠️  XML parsing error: {e}")
            return {
                'actions': 0,
                'table_count': 0,
                'tables': [],
                'referenced_scripts': set()
            }

    def _extract_tables_from_xml(self, xml_text: str) -> List[str]:
        """Extract table names from XML"""
        tables = set()
        
        # HDFS paths
        hdfs_patterns = re.findall(r'hdfs://[^<>"\s]+', xml_text)
        for path in hdfs_patterns:
            segments = path.split('/')
            skip_words = ['warehouse', 'user', 'tmp', 'stage', 'staging', 'temp', 'hdfs', 'nameservice']
            if not any(word in segments for word in skip_words):
                for seg in reversed(segments):
                    if seg and len(seg) > 1 and not seg.startswith('$'):
                        tables.add(seg)
                        break
        
        # Table properties
        table_patterns = re.findall(r'<name>table</name>\s*<value>([^<]+)</value>', xml_text)
        tables.update(table_patterns)
        
        return list(tables)

    def _match_scripts(self, referenced: Set[str], inventory: Dict) -> Tuple[Dict, Set]:
        """Match referenced scripts with inventory"""
        matched = {}
        missing = set()
        
        for script_name in referenced:
            found = False
            
            # Search in all script types
            for script_type, scripts in inventory.items():
                for script in scripts:
                    if script['name'] == script_name:
                        matched[script_name] = script
                        found = True
                        break
                if found:
                    break
            
            if not found:
                missing.add(script_name)
        
        return matched, missing

    def _mark_script_used(self, script_name: str, inventory: Dict):
        """Mark a script as used in inventory"""
        for script_type, scripts in inventory.items():
            for script in scripts:
                if script['name'] == script_name:
                    script['used'] = True
                    return

    def _count_lines(self, file_path: Path) -> int:
        """Count lines in a file"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return len(f.readlines())
        except:
            return 0

    def _classify_pipeline_length(self, scripts: int, loc: int, actions: int, tables: int) -> str:
        """Classify pipeline size based on number of scripts"""
        # Primary classification based on script count
        if scripts <= 2:
            return "Small"
        elif scripts <= 6:
            return "Medium"
        else:
            return "Big Pipeline"
    
    def _estimate_effort(self, scripts: int, loc: int, actions: int, tables: int, pipeline_size: str) -> dict:
        """
        Estimate effort for Hadoop pipeline analysis based on:
        - Discovery & Analysis (10 actions per workflow)
        - Code Reading
        - Mapping Creation  
        - Documentation
        - Validation & Review
        - Comparison with Databricks
        
        Returns: dict with hours, days, and breakdown
        """
        # Base effort per component (in hours)
        BASE_DISCOVERY = 2  # Workflow understanding
        BASE_CODE_READING = 0.5  # Per script
        BASE_MAPPING = 1  # Per script for STTM
        BASE_DOCUMENTATION = 1  # Per pipeline
        BASE_VALIDATION = 1  # Per pipeline
        BASE_COMPARISON = 2  # Compare with Databricks
        
        # Calculate effort by component
        discovery_hours = BASE_DISCOVERY + (actions * 0.2)  # +0.2h per action
        code_reading_hours = scripts * BASE_CODE_READING * (1 + loc/1000)  # Adjusted by LOC
        mapping_hours = scripts * BASE_MAPPING * (1 + tables * 0.1)  # More tables = more mapping
        documentation_hours = BASE_DOCUMENTATION
        validation_hours = BASE_VALIDATION + (scripts * 0.3)
        comparison_hours = BASE_COMPARISON
        
        # Total effort
        total_hours = (
            discovery_hours + 
            code_reading_hours + 
            mapping_hours + 
            documentation_hours + 
            validation_hours + 
            comparison_hours
        )
        
        # Apply complexity multiplier based on pipeline size
        if pipeline_size == "Big Pipeline":
            total_hours *= 1.3  # 30% overhead for complexity
        elif pipeline_size == "Medium":
            total_hours *= 1.1  # 10% overhead
        
        # Convert to days (assuming 8 hours per day)
        total_days = total_hours / 8
        
        # Round to high-level numbers (0.5 increments for cleaner estimates)
        rounded_days = round(total_days * 2) / 2  # Round to nearest 0.5
        rounded_hours = rounded_days * 8
        
        # Create effort breakdown
        effort_breakdown = {
            'total_hours': rounded_hours,
            'total_days': rounded_days,
            'discovery_hours': round(discovery_hours, 1),
            'code_reading_hours': round(code_reading_hours, 1),
            'mapping_hours': round(mapping_hours, 1),
            'documentation_hours': round(documentation_hours, 1),
            'validation_hours': round(validation_hours, 1),
            'comparison_hours': round(comparison_hours, 1),
            'effort_estimate': self._format_effort_estimate(rounded_hours, rounded_days)
        }
        
        return effort_breakdown
    
    def _format_effort_estimate(self, hours: float, days: float) -> str:
        """Format effort estimate as human-readable string"""
        # Format as clean integers or .5 increments
        if days % 1 == 0:
            days_str = f"{int(days)}"
        else:
            days_str = f"{days:.1f}"
        
        if hours % 1 == 0:
            hours_str = f"{int(hours)}"
        else:
            hours_str = f"{hours:.1f}"
        
        if days < 1:
            return f"{hours_str} hours"
        elif days < 2:
            return f"{days_str} day ({hours_str} hours)"
        else:
            return f"{days_str} days ({hours_str} hours)"

    def export_to_excel(self, output_file: str = None):
        """Export to Excel with comprehensive analysis"""
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"Hadoop_Complete_Analysis_{timestamp}.xlsx"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()

            # ==================== SHEET 1: Pipeline Summary ====================
            ws1 = wb.active
            ws1.title = 'Pipeline Summary'

            headers1 = ['Repo', 'Pipeline Name', 'Module', 'No. of Scripts', 'Pipeline Length', 'Actions', 'Tables', 'LOC', 'Scripts Missing', 'Estimated Effort']
            ws1.append(headers1)

            # Style headers
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws1[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # Add data
            for r in self.repos_data:
                ws1.append([
                    r['repo'],
                    r['pipeline_name'],
                    r['module'],
                    r['no_of_scripts'],
                    r['pipeline_length'],
                    r['actions'],
                    r['tables'],
                    r['total_loc'],
                    r['scripts_missing'],
                    r['effort_estimate']
                ])

            # Apply formatting
            for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=len(headers1)):
                for cell in row:
                    cell.border = thin_border
                    if cell.column in [4, 6, 7, 8, 9]:
                        cell.alignment = Alignment(horizontal='center')

            # Set column widths
            ws1.column_dimensions['A'].width = 25  # Repo
            ws1.column_dimensions['B'].width = 40  # Pipeline Name
            ws1.column_dimensions['C'].width = 25  # Module
            ws1.column_dimensions['D'].width = 15  # No. of Scripts
            ws1.column_dimensions['E'].width = 18  # Pipeline Length
            ws1.column_dimensions['F'].width = 10  # Actions
            ws1.column_dimensions['G'].width = 10  # Tables
            ws1.column_dimensions['H'].width = 12  # LOC
            ws1.column_dimensions['I'].width = 15  # Scripts Missing
            ws1.column_dimensions['J'].width = 20  # Estimated Effort

            # Add summary
            summary_start_row = ws1.max_row + 3
            
            total_pipelines = len(self.repos_data)
            big_count = sum(1 for r in self.repos_data if r['pipeline_length'] == 'Big Pipeline')
            medium_count = sum(1 for r in self.repos_data if r['pipeline_length'] == 'Medium')
            small_count = sum(1 for r in self.repos_data if r['pipeline_length'] == 'Small')
            total_scripts = sum(r['no_of_scripts'] for r in self.repos_data)
            total_missing = sum(r['scripts_missing'] for r in self.repos_data)
            total_effort_days = sum(r['effort_days'] for r in self.repos_data)
            total_effort_hours = sum(r['effort_hours'] for r in self.repos_data)
            
            summary_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            summary_font = Font(bold=True, size=10)
            
            ws1[f'H{summary_start_row}'] = 'Pipeline Summary'
            ws1[f'H{summary_start_row}'].font = Font(bold=True, size=12)
            ws1.merge_cells(f'H{summary_start_row}:I{summary_start_row}')
            
            summary_data = [
                ['Big Pipelines', big_count],
                ['Medium', medium_count],
                ['Small', small_count],
                ['Total Pipelines', total_pipelines],
                ['Total Scripts', total_scripts],
                ['Scripts Missing', total_missing],
                ['Total Effort (Days)', f"{int(total_effort_days) if total_effort_days % 1 == 0 else total_effort_days:.1f} days"],
                ['Total Effort (Hours)', f"{int(total_effort_hours) if total_effort_hours % 1 == 0 else total_effort_hours:.1f} hours"]
            ]
            
            for i, (label, value) in enumerate(summary_data, start=1):
                row_num = summary_start_row + i
                ws1[f'H{row_num}'] = label
                ws1[f'I{row_num}'] = value
                ws1[f'H{row_num}'].fill = summary_fill
                ws1[f'I{row_num}'].fill = summary_fill
                ws1[f'H{row_num}'].font = summary_font
                ws1[f'I{row_num}'].font = Font(bold=True, size=10, color='0000FF')
                ws1[f'H{row_num}'].border = thin_border
                ws1[f'I{row_num}'].border = thin_border

            # ==================== SHEET 2: Script Details ====================
            ws2 = wb.create_sheet('Script Details')
            
            headers2 = ['Repo', 'Pipeline Name', 'Module', 'Script Name', 'Type', 'Lines', 'Status']
            ws2.append(headers2)
            
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            for r in self.repos_data:
                # Found scripts
                for script_name, script_data in r['matched_scripts'].items():
                    # Determine type from inventory
                    script_type = 'Unknown'
                    for stype, scripts in self.script_inventory.get(r['repo'], {}).items():
                        for s in scripts:
                            if s['name'] == script_name:
                                script_type = stype
                                break
                    
                    ws2.append([
                        r['repo'],
                        r['pipeline_name'],
                        r['module'],
                        script_name,
                        script_type,
                        script_data.get('lines', 0),
                        '✓ Found'
                    ])
                
                # Missing scripts
                for script_name in r['missing_scripts']:
                    ws2.append([
                        r['repo'],
                        r['pipeline_name'],
                        r['module'],
                        script_name,
                        'Unknown',
                        0,
                        '✗ Missing'
                    ])
            
            # Apply formatting
            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=len(headers2)):
                for cell in row:
                    cell.border = thin_border
                    if cell.column == 6:
                        cell.alignment = Alignment(horizontal='center')
                    if cell.column == 7:
                        if '✗' in str(cell.value):
                            cell.font = Font(color='FF0000', bold=True)
                        else:
                            cell.font = Font(color='008000')
            
            ws2.column_dimensions['A'].width = 25
            ws2.column_dimensions['B'].width = 40
            ws2.column_dimensions['C'].width = 25
            ws2.column_dimensions['D'].width = 40
            ws2.column_dimensions['E'].width = 15
            ws2.column_dimensions['F'].width = 10
            ws2.column_dimensions['G'].width = 15

            # ==================== SHEET 3: Complete Script Inventory ====================
            ws3 = wb.create_sheet('Script Inventory')
            
            headers3 = ['Repo', 'Script Name', 'Type', 'Path', 'Lines', 'Used in Workflow']
            ws3.append(headers3)
            
            for cell in ws3[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Add all scripts from inventory
            for repo_name, inventory in self.script_inventory.items():
                for script_type, scripts in inventory.items():
                    for script in scripts:
                        ws3.append([
                            repo_name,
                            script['name'],
                            script_type,
                            script['relative_path'],
                            script['lines'],
                            '✓ Yes' if script['used'] else '✗ No'
                        ])
            
            # Apply formatting
            for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=len(headers3)):
                for cell in row:
                    cell.border = thin_border
                    if cell.column == 5:
                        cell.alignment = Alignment(horizontal='center')
                    if cell.column == 6:
                        if '✗' in str(cell.value):
                            cell.font = Font(color='FF6600')
                        else:
                            cell.font = Font(color='008000')
            
            ws3.column_dimensions['A'].width = 25
            ws3.column_dimensions['B'].width = 40
            ws3.column_dimensions['C'].width = 15
            ws3.column_dimensions['D'].width = 60
            ws3.column_dimensions['E'].width = 10
            ws3.column_dimensions['F'].width = 18

            # ==================== SHEET 4: Effort Breakdown ====================
            ws4 = wb.create_sheet('Effort Breakdown')
            
            headers4 = ['Repo', 'Pipeline Name', 'Module', 'Pipeline Length', 
                       'Total Days', 'Total Hours', 
                       'Discovery', 'Code Reading', 'Mapping', 'Documentation', 'Validation', 'Comparison']
            ws4.append(headers4)
            
            for cell in ws4[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            for r in self.repos_data:
                effort = r['effort_breakdown']
                ws4.append([
                    r['repo'],
                    r['pipeline_name'],
                    r['module'],
                    r['pipeline_length'],
                    effort['total_days'],
                    effort['total_hours'],
                    effort['discovery_hours'],
                    effort['code_reading_hours'],
                    effort['mapping_hours'],
                    effort['documentation_hours'],
                    effort['validation_hours'],
                    effort['comparison_hours']
                ])
            
            # Apply formatting
            for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, min_col=1, max_col=len(headers4)):
                for cell in row:
                    cell.border = thin_border
                    if cell.column in [5, 6, 7, 8, 9, 10, 11, 12]:
                        cell.alignment = Alignment(horizontal='center')
                        if cell.column in [5, 6]:
                            cell.font = Font(bold=True, color='0000FF')
            
            ws4.column_dimensions['A'].width = 25
            ws4.column_dimensions['B'].width = 40
            ws4.column_dimensions['C'].width = 25
            ws4.column_dimensions['D'].width = 18
            ws4.column_dimensions['E'].width = 12
            ws4.column_dimensions['F'].width = 12
            ws4.column_dimensions['G'].width = 12
            ws4.column_dimensions['H'].width = 14
            ws4.column_dimensions['I'].width = 12
            ws4.column_dimensions['J'].width = 15
            ws4.column_dimensions['K'].width = 12
            ws4.column_dimensions['L'].width = 12

            # Save
            wb.save(output_file)
            print(f"\n✅ Excel report generated: {output_file}")
            print(f"   📊 {total_pipelines} pipelines analyzed")
            print(f"   📝 {total_scripts} scripts referenced ({total_scripts - total_missing} found, {total_missing} missing)")
            
            # Format effort cleanly
            days_display = int(total_effort_days) if total_effort_days % 1 == 0 else f"{total_effort_days:.1f}"
            hours_display = int(total_effort_hours) if total_effort_hours % 1 == 0 else f"{total_effort_hours:.1f}"
            print(f"   ⏱️  Total effort: {days_display} days ({hours_display} hours)")
            print(f"   📦 Pipeline breakdown: {big_count} Big, {medium_count} Medium, {small_count} Small")
            
            # Print inventory summary
            print(f"\n📦 Script Inventory Summary:")
            for repo_name, inventory in self.script_inventory.items():
                total = sum(len(scripts) for scripts in inventory.values())
                used = sum(1 for scripts in inventory.values() for s in scripts if s['used'])
                print(f"   {repo_name}: {total} scripts ({used} used, {total-used} unused)")

        except Exception as e:
            print(f"⚠️  Excel generation error: {e}")
            import traceback
            traceback.print_exc()


def main():
    """Main function"""
    import argparse

    parser = argparse.ArgumentParser(
        description='Complete Hadoop Repository Analyzer - Two-Phase Analysis'
    )

    parser.add_argument(
        'path',
        help='Path to Hadoop repository or folder containing multiple repos'
    )

    parser.add_argument(
        '--output',
        type=str,
        default=None,
        help='Output Excel file name'
    )

    args = parser.parse_args()

    analyzer = CompleteHadoopAnalyzer()

    path = Path(args.path)
    if path.is_dir():
        # Check if it contains multiple repos (subdirectories that are repos)
        subdirs = [item for item in path.iterdir() if item.is_dir() and not item.name.startswith('.')]
        sub_repos = [d for d in subdirs if analyzer._is_hadoop_repo(d)]
        
        if len(sub_repos) > 0:
            # Multi-repo mode: folder contains multiple repos
            print("📦 Multi-repository mode\n")
            analyzer.analyze_folder(args.path)
        elif analyzer._is_hadoop_repo(path):
            # Single repo mode: the path itself is a repo
            print("📦 Single repository mode\n")
            analyzer.analyze_repo(args.path)
        else:
            print(f"⚠️  No Hadoop repositories found in: {args.path}")
            return
    else:
        print(f"⚠️  Invalid path: {args.path}")
        return

    # Export
    analyzer.export_to_excel(args.output)

    print("\n" + "="*100)
    print("✅ Analysis Complete!")
    print("="*100)


if __name__ == "__main__":
    main()

